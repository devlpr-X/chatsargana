import uuid
import csv
import io
import calendar
from datetime import date, datetime
from decimal import Decimal
from functools import wraps

from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.mail import send_mail, EmailMessage
from django.conf import settings
from django.db.models import Sum, Count, F, DecimalField
from django.utils import timezone
from django.views.decorators.http import require_POST

from .models import (
    SiteContent, HeroStat, Feature, AboutCard, BenefitTeaser,
    ProductCategory, Product,
    NutrientCard, UsageCard,
    KeyNumber, BenefitDetail,
    MissionValue, ProcessStep,
    Order, OrderItem,
    PasswordResetToken,
)


def sc(key, default=''):
    """Get SiteContent value by key."""
    try:
        return SiteContent.objects.get(key=key).value
    except SiteContent.DoesNotExist:
        return default


# ─────────────────────────────────────────────────────────────
#  Cart helpers
# ─────────────────────────────────────────────────────────────
def _get_cart(request):
    return request.session.get('cart', {})


def _save_cart(request, cart):
    request.session['cart'] = cart
    request.session.modified = True


def _cart_items(cart):
    items = []
    total = Decimal('0')
    for pid, item in cart.items():
        sub = Decimal(item['price']) * item['qty']
        items.append({
            'id': pid,
            'name': item['name'],
            'price': Decimal(item['price']),
            'qty': item['qty'],
            'unit': item.get('unit', ''),
            'image': item.get('image', ''),
            'subtotal': sub,
        })
        total += sub
    return items, total


# ─────────────────────────────────────────────────────────────
#  Panel decorator
# ─────────────────────────────────────────────────────────────
def panel_required(view_func):
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated or not request.user.is_staff:
            return redirect('/panel/login/?next=' + request.path)
        return view_func(request, *args, **kwargs)
    return _wrapped


# ─────────────────────────────────────────────────────────────
#  Public pages
# ─────────────────────────────────────────────────────────────
def index(request):
    context = {
        'hero_badge':       sc('hero_badge', '100% Байгалийн · УВС аймгийн'),
        'hero_title_1':     sc('hero_title_1', 'Байгалийн'),
        'hero_title_2':     sc('hero_title_2', 'алт жимс'),
        'hero_subtitle':    sc('hero_subtitle', 'Увсын тал хээрийн цэвэр агаарт өссөн чацарганы жимснээс гарган авсан шим тэжээлт бүтээгдэхүүнүүд — таны эрүүл мэндийн найдвартай түнш.'),
        'hero_stats':       HeroStat.objects.all(),
        'features':         Feature.objects.all(),
        'about_label':      sc('about_teaser_label', 'Чацаргана гэж юу вэ'),
        'about_title':      sc('about_teaser_title', 'Байгалийн <em>нандин</em><br>жимс'),
        'about_body':       sc('about_teaser_body', 'Чацаргана нь тод улбар шар өнгөтэй, хүчиллэг амттай жимс бөгөөд дэлхийд хамгийн их шим тэжээл агуулдаг жимсний нэгд тооцогддог.'),
        'about_cards':      AboutCard.objects.all(),
        'preview_products': Product.objects.filter(is_active=True).order_by('order')[:4],
        'benefit_teasers':  BenefitTeaser.objects.all(),
        'products_label':   sc('products_label', 'Манай бүтээгдэхүүнүүд'),
        'products_title':   sc('products_title', 'Чацарганы <em>цуглуулга</em>'),
        'benefits_label':   sc('benefits_label', 'Ашиг тус'),
        'benefits_title':   sc('benefits_title', 'Эрүүл мэндэд <em>үзүүлэх нөлөө</em>'),
    }
    return render(request, 'index.html', context)


def info(request):
    context = {
        'hero_subtitle':    sc('info_hero_subtitle', 'Байгалийн хамгийн баялаг шим тэжээлт жимсний нэгийн тухай'),
        'story_label':      sc('info_story_label', 'Гарал үүсэл'),
        'story_title_1':    sc('info_story_title_1', 'Монголын нутгийн'),
        'story_title_2':    sc('info_story_title_2', 'алтан жимс'),
        'story_body_1':     sc('info_story_body_1', 'Чацаргана <em>(Hippophae rhamnoides)</em> нь тод улбар шар өнгөтэй жимс.'),
        'story_body_2':     sc('info_story_body_2', 'УВС аймгийн нутагт ургасан чацаргана нь цэвэр агаарт бойжсон.'),
        'story_body_3':     sc('info_story_body_3', 'Уламжлалт монгол эмнэлэгт олон зуун жилийн турш хэрэглэж ирсэн.'),
        'fact_num':         sc('info_fact_num', '1956'),
        'fact_text':        sc('info_fact_text', 'он — Оросын сансрын аялагчдын хоолонд анх оруулсан'),
        'nutrition_sub':    sc('info_nutrition_sub', '100 грамм чацарганы жимсэнд агуулагдах гол бодисууд'),
        'nutrient_cards':   NutrientCard.objects.all(),
        'usage_cards':      UsageCard.objects.all(),
        'cta_body':         sc('info_cta_body', 'Байгалийн чацарганаас гарган авсан дэлгэрэнгүй бүтээгдэхүүнүүдтэй танилцаарай'),
    }
    return render(request, 'info.html', context)


def products(request):
    categories = ProductCategory.objects.all()
    all_products = Product.objects.filter(is_active=True).select_related('category').order_by('order')
    context = {
        'categories': categories,
        'products':   all_products,
    }
    return render(request, 'products.html', context)


def benefits(request):
    context = {
        'hero_sub':    sc('angiinfo_hero_sub', 'Эрдэм шинжилгээгээр нотлогдсон байгалийн нандин жимсний биед үзүүлэх нөлөө.'),
        'key_numbers': KeyNumber.objects.all(),
        'benefits':    BenefitDetail.objects.all(),
        'cta_body':    sc('angiinfo_cta_body', 'Чацарганы бүтээгдэхүүнүүдтэй танилцаж, өөрийнхөө хэрэгцээнд тохирохыг сонгоорой'),
    }
    return render(request, 'angiinfo.html', context)


def about(request):
    context = {
        'hero_sub':        sc('about_hero_sub', 'Монголын баялаг байгалийг хайрлаж, Увсын тал нутгийн чацарганыг дэлхийд таниулах зорилготой.'),
        'mission_body_1':  sc('about_mission_body_1', 'Бид Увс аймгийн байгалийн нөхцөлд ургасан чацарганыг хиймэл нэмэлтгүй боловсруулан хүргэдэг.'),
        'mission_body_2':  sc('about_mission_body_2', 'Манай бүтээгдэхүүн бүр хатуу чанарын шалгуур давдаг.'),
        'mission_values':  MissionValue.objects.all(),
        'card_big_text':   sc('about_card_big_text', 'УВС аймгийн тал хээрийн байгалиас цуглуулсан чацарганы жимс.'),
        'card_year':       sc('about_card_year', '2025'),
        'process_steps':   ProcessStep.objects.all(),
        'contact_address': sc('contact_address', 'УВС аймаг, Монгол улс'),
        'contact_phone':   sc('contact_phone', '+976 ···· ····'),
        'contact_email':   sc('contact_email', 'info@chatsargana.mn'),
        'contact_social':  sc('contact_social', 'Facebook · Instagram'),
    }
    return render(request, 'about.html', context)


# ─────────────────────────────────────────────────────────────
#  Cart views
# ─────────────────────────────────────────────────────────────
def cart_view(request):
    cart = _get_cart(request)
    items, total = _cart_items(cart)
    return render(request, 'cart.html', {'items': items, 'total': total})


@require_POST
def cart_add(request, product_id):
    try:
        product = Product.objects.get(pk=product_id, is_active=True)
    except Product.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Бүтээгдэхүүн олдсонгүй'}, status=404)

    qty = max(1, int(request.POST.get('qty', 1)))
    cart = _get_cart(request)
    pid = str(product_id)

    if pid in cart:
        cart[pid]['qty'] += qty
    else:
        cart[pid] = {
            'qty':   qty,
            'price': str(product.price),
            'name':  product.name,
            'unit':  product.unit,
            'image': product.image.url if product.image else '',
        }
    _save_cart(request, cart)

    count = sum(i['qty'] for i in cart.values())
    return JsonResponse({'ok': True, 'cart_count': count})


@require_POST
def cart_remove(request, product_id):
    cart = _get_cart(request)
    cart.pop(str(product_id), None)
    _save_cart(request, cart)

    count = sum(i['qty'] for i in cart.values())
    total = sum(Decimal(i['price']) * i['qty'] for i in cart.values())
    return JsonResponse({'ok': True, 'cart_count': count, 'cart_total': str(total)})


@require_POST
def cart_update(request, product_id):
    cart = _get_cart(request)
    pid = str(product_id)
    qty = int(request.POST.get('qty', 1))

    if pid in cart:
        if qty <= 0:
            cart.pop(pid)
        else:
            cart[pid]['qty'] = qty
    _save_cart(request, cart)

    count   = sum(i['qty'] for i in cart.values())
    total   = sum(Decimal(i['price']) * i['qty'] for i in cart.values())
    subtotal = Decimal(cart[pid]['price']) * qty if (pid in cart and qty > 0) else Decimal('0')
    return JsonResponse({
        'ok':         True,
        'cart_count': count,
        'cart_total': str(total),
        'subtotal':   str(subtotal),
    })


# ─────────────────────────────────────────────────────────────
#  Checkout & Orders
# ─────────────────────────────────────────────────────────────
def checkout(request):
    # Админ хэрэглэгч захиалга өгдөггүй — зөвхөн хяналт хийнэ
    if request.user.is_authenticated and request.user.is_staff:
        messages.info(request, 'Админ хэрэглэгч захиалга өгөх боломжгүй — зөвхөн захиалга хянана.')
        return redirect('core:panel_dashboard')

    cart = _get_cart(request)
    if not cart:
        return redirect('core:cart')

    items, total = _cart_items(cart)

    if request.method == 'POST':
        name    = request.POST.get('name', '').strip()
        phone   = request.POST.get('phone', '').strip()
        email   = request.POST.get('email', '').strip()
        address = request.POST.get('address', '').strip()
        lat     = request.POST.get('latitude', '').strip()
        lng     = request.POST.get('longitude', '').strip()
        note    = request.POST.get('note', '').strip()

        errors = {}
        if not name:    errors['name']    = 'Нэр оруулна уу'
        if not phone:   errors['phone']   = 'Утасны дугаар оруулна уу'
        if not address: errors['address'] = 'Хаяг оруулна уу'

        if not errors:
            if not request.session.session_key:
                request.session.create()

            order = Order.objects.create(
                user           = request.user if request.user.is_authenticated else None,
                session_key    = request.session.session_key or '',
                customer_name  = name,
                customer_phone = phone,
                customer_email = email,
                address        = address,
                latitude       = float(lat) if lat else None,
                longitude      = float(lng) if lng else None,
                note           = note,
                total_price    = total,
            )
            for pid, item in cart.items():
                product = None
                try:
                    product = Product.objects.get(pk=int(pid))
                except (Product.DoesNotExist, ValueError):
                    pass
                OrderItem.objects.create(
                    order        = order,
                    product      = product,
                    product_name = item['name'],
                    quantity     = item['qty'],
                    price        = Decimal(item['price']),
                )

            _notify_admin_new_order(request, order)
            _save_cart(request, {})
            return redirect('core:order_confirm', order_id=order.pk)

        return render(request, 'checkout.html', {
            'items':     items,
            'total':     total,
            'errors':    errors,
            'form_data': request.POST,
        })

    form_data = {}
    if request.user.is_authenticated:
        form_data = {
            'name':  request.user.get_full_name() or request.user.username,
            'email': request.user.email,
        }
    return render(request, 'checkout.html', {
        'items':     items,
        'total':     total,
        'form_data': form_data,
    })


def order_confirm(request, order_id):
    order = get_object_or_404(Order, pk=order_id)
    # Security: only the owner (by user or session) or staff can view
    if not request.user.is_staff:
        if order.user:
            if not request.user.is_authenticated or order.user != request.user:
                return redirect('core:index')
        else:
            if order.session_key != (request.session.session_key or ''):
                return redirect('core:index')
    return render(request, 'order_confirm.html', {'order': order})


@login_required(login_url='/panel/login/')
def my_orders(request):
    orders = Order.objects.filter(user=request.user).prefetch_related('items').order_by('-created_at')
    return render(request, 'my_orders.html', {'orders': orders})


# ─────────────────────────────────────────────────────────────
#  Admin Panel
# ─────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────
#  User auth (login / logout / register)
# ─────────────────────────────────────────────────────────────
def user_login(request):
    if request.user.is_authenticated:
        return redirect('core:index')
    error = None
    if request.method == 'POST':
        email = request.POST.get('username', '').strip().lower()
        password = request.POST.get('password', '')
        try:
            user_obj = User.objects.get(email__iexact=email)
            user = authenticate(request, username=user_obj.username, password=password)
        except User.DoesNotExist:
            user = None
        if user:
            login(request, user)
            return redirect(request.GET.get('next', '/'))
        error = 'И-мэйл эсвэл нууц үг буруу байна'
    return render(request, 'user_login.html', {'error': error})


def user_logout(request):
    logout(request)
    return redirect('core:index')


def password_reset_request(request):
    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        try:
            user = User.objects.get(email__iexact=email)
        except User.DoesNotExist:
            # Аюулгүй байдлын үүднээс и-мэйл байгаа эсэхийг нуух
            return render(request, 'password_reset_request.html', {'sent': True})

        # Хуучин токенуудыг хүчингүй болгох
        PasswordResetToken.objects.filter(user=user, used=False).update(used=True)

        token = PasswordResetToken.objects.create(user=user)
        reset_url = request.build_absolute_uri(f'/reset-password/confirm/?token={token.token}')

        try:
            send_mail(
                subject='Нууц үг сэргээх — Чацаргана',
                message=(
                    f'Сайн байна уу!\n\n'
                    f'Та доорх холбоосоор орж нууц үгээ шинэчлэх боломжтой:\n\n'
                    f'{reset_url}\n\n'
                    f'Энэ холбоос 1 цагийн дотор хүчинтэй.\n\n'
                    f'Хэрэв та энэ хүсэлтийг гаргаагүй бол энэ мэйлийг үл тоомсорлоно уу.\n\n'
                    f'— Чацаргана баг'
                ),
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=False,
            )
        except Exception:
            return render(request, 'password_reset_request.html', {
                'error': 'И-мэйл илгээхэд алдаа гарлаа. Дахин оролдоно уу.'
            })

        return render(request, 'password_reset_request.html', {'sent': True})

    return render(request, 'password_reset_request.html', {})


def password_reset_confirm(request):
    token_str = request.GET.get('token') or request.POST.get('token', '')
    error = None
    token_obj = None

    try:
        token_obj = PasswordResetToken.objects.select_related('user').get(
            token=uuid.UUID(str(token_str))
        )
        if not token_obj.is_valid():
            error = 'Холбоосын хугацаа дууссан эсвэл ашигласан байна. Дахин хүсэлт илгээнэ үү.'
            token_obj = None
    except (PasswordResetToken.DoesNotExist, ValueError, TypeError, AttributeError):
        error = 'Буруу эсвэл хүчингүй холбоос байна.'

    if request.method == 'POST' and token_obj:
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')

        if not password1:
            error = 'Нууц үг оруулна уу'
        elif len(password1) < 6:
            error = 'Нууц үг хамгийн багадаа 6 тэмдэгт байх ёстой'
        elif password1 != password2:
            error = 'Нууц үг таарахгүй байна'
        else:
            user = token_obj.user
            user.set_password(password1)
            user.save()
            token_obj.used = True
            token_obj.save()
            return render(request, 'password_reset_confirm.html', {'success': True})

    return render(request, 'password_reset_confirm.html', {
        'token':       token_str,
        'token_valid': token_obj is not None,
        'error':       error,
    })


def user_register(request):
    if request.user.is_authenticated:
        return redirect('core:index')
    errors = {}
    if request.method == 'POST':
        email     = request.POST.get('email', '').strip()
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')

        if not email:                        errors['email']     = 'И-мэйл оруулна уу'
        elif User.objects.filter(email=email).exists():
                                             errors['email']     = 'Энэ и-мэйл бүртгэлтэй байна'
        if not password1:                    errors['password1'] = 'Нууц үг оруулна уу'
        elif len(password1) < 6:             errors['password1'] = 'Нууц үг хамгийн багадаа 6 тэмдэгт'
        elif password1 != password2:         errors['password2'] = 'Нууц үг таарахгүй байна'

        if not errors:
            username = email.split('@')[0]
            base, i = username, 1
            while User.objects.filter(username=username).exists():
                username = f'{base}{i}'; i += 1
            user = User.objects.create_user(username=username, password=password1, email=email)
            login(request, user)
            return redirect('core:index')
    return render(request, 'user_register.html', {'errors': errors, 'form_data': request.POST})


def panel_login_view(request):
    if request.user.is_authenticated and request.user.is_staff:
        return redirect('core:panel_dashboard')

    error = None
    if request.method == 'POST':
        username = request.POST.get('username', '')
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user and user.is_staff:
            login(request, user)
            return redirect(request.GET.get('next', '/panel/'))
        error = 'Нэвтрэх нэр эсвэл нууц үг буруу байна'

    return render(request, 'panel_login.html', {'error': error})


def panel_logout_view(request):
    logout(request)
    return redirect('core:panel_login')


@panel_required
def panel_dashboard(request):
    stats = {s[0]: Order.objects.filter(status=s[0]).count() for s in Order.STATUS_CHOICES}
    total_revenue = Order.objects.filter(
        status__in=['confirmed', 'shipped', 'delivered']
    ).aggregate(t=Sum('total_price'))['t'] or 0
    recent_orders = Order.objects.select_related('user').prefetch_related('items')[:8]

    context = {
        'stats':         stats,
        'total_revenue': total_revenue,
        'recent_orders': recent_orders,
        'status_choices': Order.STATUS_CHOICES,
    }
    return render(request, 'panel_dashboard.html', context)


@panel_required
def panel_orders(request):
    status_filter = request.GET.get('status', '')
    qs = Order.objects.select_related('user').prefetch_related('items')
    if status_filter:
        qs = qs.filter(status=status_filter)
    context = {
        'orders':         qs,
        'status_filter':  status_filter,
        'status_choices': Order.STATUS_CHOICES,
        'status_colors':  Order.STATUS_COLORS,
    }
    return render(request, 'panel_orders.html', context)


# ─────────────────────────────────────────────────────────────
#  Admin email notification (new order)
# ─────────────────────────────────────────────────────────────
def _notify_admin_new_order(request, order):
    """Шинэ захиалга үүсэхэд админы и-мэйл рүү мэдэгдэл илгээнэ."""
    admin_email = getattr(settings, 'ADMIN_NOTIFY_EMAIL', '') or settings.DEFAULT_FROM_EMAIL
    if not admin_email:
        return
    try:
        panel_url = request.build_absolute_uri(
            reverse('core:panel_order_detail', args=[order.pk])
        )
    except Exception:
        panel_url = ''

    lines = [
        f'Шинэ захиалга #{order.pk} ирлээ.',
        '',
        f'Нэр:   {order.customer_name}',
        f'Утас:  {order.customer_phone}',
        f'И-мэйл: {order.customer_email or "—"}',
        f'Хаяг:  {order.address}',
    ]
    if order.note:
        lines.append(f'Тэмдэглэл: {order.note}')
    lines.append('')
    lines.append('Захиалсан бараа:')
    for it in order.items.all():
        lines.append(f'  • {it.product_name} × {it.quantity} = ₮{it.subtotal:,.0f}')
    lines.append('')
    lines.append(f'Нийт дүн: ₮{order.total_price:,.0f}')
    if panel_url:
        lines.append('')
        lines.append(f'Дэлгэрэнгүй: {panel_url}')

    try:
        send_mail(
            subject=f'🛒 Шинэ захиалга #{order.pk} — {order.customer_name}',
            message='\n'.join(lines),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[admin_email],
            fail_silently=True,
        )
    except Exception:
        pass


# ─────────────────────────────────────────────────────────────
#  Reports — date presets + aggregation
# ─────────────────────────────────────────────────────────────
PRESET_LABELS = [
    ('this_month',    'Энэ сар'),
    ('last_month',    'Өмнөх сар'),
    ('last_3_months', 'Сүүлийн 3 сар'),
    ('last_6_months', 'Сүүлийн 6 сар'),
    ('this_year',     'Энэ жил'),
    ('last_year',     'Өмнөх жил'),
    ('all',           'Бүх хугацаа'),
]


def _month_bounds(year, month):
    first = date(year, month, 1)
    last = date(year, month, calendar.monthrange(year, month)[1])
    return first, last


def _shift_month(year, month, delta):
    idx = year * 12 + (month - 1) + delta
    return idx // 12, idx % 12 + 1


def _preset_ranges(today):
    """Сонголт бүрийн (эхлэх, дуусах) огноог буцаана."""
    y, m = today.year, today.month
    ranges = {}
    ranges['this_month'] = _month_bounds(y, m)
    ly, lm = _shift_month(y, m, -1)
    ranges['last_month'] = _month_bounds(ly, lm)
    s3y, s3m = _shift_month(y, m, -2)
    ranges['last_3_months'] = (_month_bounds(s3y, s3m)[0], _month_bounds(y, m)[1])
    s6y, s6m = _shift_month(y, m, -5)
    ranges['last_6_months'] = (_month_bounds(s6y, s6m)[0], _month_bounds(y, m)[1])
    ranges['this_year'] = (date(y, 1, 1), date(y, 12, 31))
    ranges['last_year'] = (date(y - 1, 1, 1), date(y - 1, 12, 31))
    first_dt = Order.objects.order_by('created_at').values_list('created_at', flat=True).first()
    start_all = timezone.localtime(first_dt).date() if first_dt else date(y, 1, 1)
    ranges['all'] = (start_all, today)
    return ranges


def _resolve_range(request):
    """GET-ээс preset эсвэл custom огноог задлаж (preset, start, end) буцаана."""
    today = timezone.localdate()
    ranges = _preset_ranges(today)
    preset = request.GET.get('preset', 'this_month')
    if preset not in ranges:
        preset = 'this_month'

    start_str = request.GET.get('start', '').strip()
    end_str = request.GET.get('end', '').strip()
    if start_str and end_str:
        try:
            start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
            if start_date > end_date:
                start_date, end_date = end_date, start_date
            return 'custom', start_date, end_date, ranges
        except ValueError:
            pass
    start_date, end_date = ranges[preset]
    return preset, start_date, end_date, ranges


def _report_data(start_date, end_date):
    """Тухайн хугацааны нэгтгэсэн тайлангийн өгөгдөл."""
    orders = Order.objects.filter(
        created_at__date__gte=start_date, created_at__date__lte=end_date
    )
    non_cancelled = orders.exclude(status='cancelled')

    total_revenue = non_cancelled.aggregate(t=Sum('total_price'))['t'] or 0
    total_orders = orders.count()
    valid_orders = non_cancelled.count()

    items = OrderItem.objects.filter(order__in=non_cancelled)
    products = list(items.values('product_name').annotate(
        qty=Sum('quantity'),
        revenue=Sum(F('quantity') * F('price'),
                    output_field=DecimalField(max_digits=14, decimal_places=0)),
        order_count=Count('order', distinct=True),
    ).order_by('-qty'))

    total_items = sum(p['qty'] for p in products)
    avg_order = (total_revenue / valid_orders) if valid_orders else 0

    top5 = products[:5]
    bottom5 = sorted(products, key=lambda p: p['qty'])[:5]

    status_breakdown = []
    for val, label in Order.STATUS_CHOICES:
        status_breakdown.append({
            'value': val,
            'label': label,
            'count': orders.filter(status=val).count(),
            'color': Order.STATUS_COLORS.get(val, '#6b7280'),
        })

    return {
        'orders':           orders.select_related('user').prefetch_related('items'),
        'total_orders':     total_orders,
        'valid_orders':     valid_orders,
        'total_revenue':    total_revenue,
        'total_items':      total_items,
        'avg_order':        avg_order,
        'products':         products,
        'top5':             top5,
        'bottom5':          bottom5,
        'status_breakdown': status_breakdown,
    }


def _build_xlsx(data, start_date, end_date):
    """Тайланг .xlsx болгон bytes хэлбэрээр буцаана."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = 'Тайлан'
    bold = Font(bold=True)

    ws.append(['Чацаргана — Борлуулалтын тайлан'])
    ws['A1'].font = bold
    ws.append([f'Хугацаа: {start_date} — {end_date}'])
    ws.append([])
    ws.append(['Нийт захиалга', data['total_orders']])
    ws.append(['Хүчинтэй захиалга', data['valid_orders']])
    ws.append(['Нийт орлого (₮)', float(data['total_revenue'])])
    ws.append(['Нийт бараа (ширхэг)', int(data['total_items'])])
    ws.append(['Дундаж захиалга (₮)', float(data['avg_order'])])
    ws.append([])

    hdr = ['Бүтээгдэхүүн', 'Тоо ширхэг', 'Орлого (₮)', 'Захиалга']
    ws.append(hdr)
    for c in ws[ws.max_row]:
        c.font = bold
    for p in data['products']:
        ws.append([p['product_name'], int(p['qty']), float(p['revenue']), p['order_count']])

    for col in ('A', 'B', 'C', 'D'):
        ws.column_dimensions[col].width = 22

    # Захиалгуудын дэлгэрэнгүй sheet
    ws2 = wb.create_sheet('Захиалгууд')
    ws2.append(['#', 'Нэр', 'Утас', 'Төлөв', 'Нийт (₮)', 'Огноо'])
    for c in ws2[1]:
        c.font = bold
    for o in data['orders']:
        ws2.append([
            o.pk, o.customer_name, o.customer_phone,
            o.get_status_display(), float(o.total_price),
            timezone.localtime(o.created_at).strftime('%Y-%m-%d %H:%M'),
        ])
    for col, w in zip('ABCDEF', (8, 24, 16, 16, 14, 18)):
        ws2.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@panel_required
def panel_reports(request):
    preset, start_date, end_date, ranges = _resolve_range(request)
    data = _report_data(start_date, end_date)

    preset_json = {k: [v[0].isoformat(), v[1].isoformat()] for k, v in ranges.items()}
    chart_rows = data['products'][:10]
    chart = {
        'labels':  [p['product_name'] for p in chart_rows],
        'qty':     [int(p['qty']) for p in chart_rows],
        'revenue': [float(p['revenue']) for p in chart_rows],
    }

    context = {
        'preset':         preset,
        'preset_labels':  PRESET_LABELS,
        'preset_json':    preset_json,
        'start_date':     start_date.isoformat(),
        'end_date':       end_date.isoformat(),
        'chart':          chart,
        **data,
    }
    return render(request, 'panel_reports.html', context)


@panel_required
def panel_report_export(request, fmt):
    preset, start_date, end_date, ranges = _resolve_range(request)
    data = _report_data(start_date, end_date)
    fname = f'tailan_{start_date}_{end_date}'

    if fmt == 'csv':
        resp = HttpResponse(content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = f'attachment; filename="{fname}.csv"'
        resp.write('﻿')  # Excel-д Cyrillic зөв харагдах BOM
        w = csv.writer(resp)
        w.writerow(['Чацаргана — Борлуулалтын тайлан'])
        w.writerow([f'Хугацаа: {start_date} — {end_date}'])
        w.writerow([])
        w.writerow(['Нийт захиалга', data['total_orders']])
        w.writerow(['Нийт орлого (₮)', int(data['total_revenue'])])
        w.writerow(['Нийт бараа (ширхэг)', int(data['total_items'])])
        w.writerow([])
        w.writerow(['Бүтээгдэхүүн', 'Тоо ширхэг', 'Орлого (₮)', 'Захиалга'])
        for p in data['products']:
            w.writerow([p['product_name'], int(p['qty']), int(p['revenue']), p['order_count']])
        return resp

    if fmt in ('xlsx', 'excel'):
        content = _build_xlsx(data, start_date, end_date)
        resp = HttpResponse(
            content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="{fname}.xlsx"'
        return resp

    return JsonResponse({'ok': False, 'error': 'Unknown format'}, status=400)


@panel_required
@require_POST
def panel_report_email(request):
    preset, start_date, end_date, ranges = _resolve_range(request)
    data = _report_data(start_date, end_date)
    admin_email = getattr(settings, 'ADMIN_NOTIFY_EMAIL', '') or settings.DEFAULT_FROM_EMAIL

    if not admin_email:
        messages.error(request, 'Админы и-мэйл тохируулагдаагүй байна.')
        return redirect(f"{reverse('core:panel_reports')}?preset={preset}&start={start_date}&end={end_date}")

    lines = [
        'Чацаргана — Борлуулалтын тайлан',
        f'Хугацаа: {start_date} — {end_date}',
        '',
        f'Нийт захиалга:      {data["total_orders"]}',
        f'Хүчинтэй захиалга:  {data["valid_orders"]}',
        f'Нийт орлого:        ₮{data["total_revenue"]:,.0f}',
        f'Нийт бараа:         {data["total_items"]} ширхэг',
        f'Дундаж захиалга:    ₮{data["avg_order"]:,.0f}',
        '',
        'Их борлуулалттай 5 бараа:',
    ]
    for p in data['top5']:
        lines.append(f'  • {p["product_name"]}: {p["qty"]} ширхэг (₮{p["revenue"]:,.0f})')
    lines.append('')
    lines.append('Дэлгэрэнгүйг хавсаргасан Excel файлаас үзнэ үү.')

    try:
        email = EmailMessage(
            subject=f'📊 Борлуулалтын тайлан {start_date} — {end_date}',
            body='\n'.join(lines),
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[admin_email],
        )
        email.attach(
            f'tailan_{start_date}_{end_date}.xlsx',
            _build_xlsx(data, start_date, end_date),
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        email.send(fail_silently=False)
        messages.success(request, f'Тайлан {admin_email} хаяг руу илгээгдлээ.')
    except Exception as e:
        messages.error(request, f'И-мэйл илгээхэд алдаа гарлаа: {e}')

    return redirect(f"{reverse('core:panel_reports')}?preset={preset}&start={start_date}&end={end_date}")


# ─────────────────────────────────────────────────────────────
#  Inline edit (admin in-page editing)
# ─────────────────────────────────────────────────────────────
INLINE_EDITABLE = {
    # SiteContent is keyed by string `key`, value lives in `value`
    'sitecontent':   {'model': SiteContent,   'pk': 'key', 'fields': {'value': 'text'}},
    'herostat':      {'model': HeroStat,      'fields': {'num': 'text', 'label': 'text'}},
    'feature':       {'model': Feature,       'fields': {'icon': 'text', 'title': 'text', 'description': 'text'}},
    'aboutcard':     {'model': AboutCard,     'fields': {'value': 'text', 'label': 'text'}},
    'benefitteaser': {'model': BenefitTeaser, 'fields': {'icon': 'text', 'title': 'text', 'description': 'text'}},
    'nutrientcard':  {'model': NutrientCard,  'fields': {'letter': 'text', 'title': 'text', 'value_text': 'text', 'note': 'text', 'bar_pct': 'int'}},
    'usagecard':     {'model': UsageCard,     'fields': {'icon': 'text', 'title': 'text', 'description': 'text', 'image': 'image'}},
    'keynumber':     {'model': KeyNumber,     'fields': {'num': 'text', 'label': 'text'}},
    'benefitdetail': {'model': BenefitDetail, 'fields': {'icon': 'text', 'number': 'text', 'title': 'text', 'description': 'text', 'tags': 'text'}},
    'missionvalue':  {'model': MissionValue,  'fields': {'icon': 'text', 'title': 'text', 'description': 'text'}},
    'processstep':   {'model': ProcessStep,   'fields': {'number': 'text', 'icon': 'text', 'title': 'text', 'description': 'text'}},
    'product':       {'model': Product,       'fields': {'name': 'text', 'description': 'text', 'price': 'int', 'unit': 'text', 'image': 'image', 'badge_icon': 'text'}},
}


@require_POST
def inline_edit(request):
    if not request.user.is_authenticated or not request.user.is_staff:
        return JsonResponse({'ok': False, 'error': 'forbidden'}, status=403)

    model_name = request.POST.get('model', '').lower()
    pk        = request.POST.get('pk', '')
    field     = request.POST.get('field', '')

    spec = INLINE_EDITABLE.get(model_name)
    if not spec:
        return JsonResponse({'ok': False, 'error': f'Model "{model_name}" not allowed'}, status=400)

    field_type = spec['fields'].get(field)
    if not field_type:
        return JsonResponse({'ok': False, 'error': f'Field "{field}" not allowed'}, status=400)

    Model = spec['model']
    pk_field = spec.get('pk', 'pk')

    if pk_field == 'key':
        # SiteContent — create if missing so admins can add new content
        obj, _ = Model.objects.get_or_create(key=pk, defaults={'value': ''})
    else:
        try:
            obj = Model.objects.get(pk=pk)
        except Model.DoesNotExist:
            return JsonResponse({'ok': False, 'error': 'Object not found'}, status=404)

    if field_type == 'image':
        f = request.FILES.get('value')
        if not f:
            return JsonResponse({'ok': False, 'error': 'No image uploaded'}, status=400)
        setattr(obj, field, f)
        obj.save()
        img_field = getattr(obj, field)
        return JsonResponse({'ok': True, 'image_url': img_field.url if img_field else ''})

    value = request.POST.get('value', '')
    if field_type == 'int':
        try:
            value = int(str(value).replace(',', '').strip())
        except (ValueError, TypeError):
            return JsonResponse({'ok': False, 'error': 'Invalid integer'}, status=400)

    setattr(obj, field, value)
    obj.save()
    return JsonResponse({'ok': True, 'value': str(getattr(obj, field))})


@panel_required
def panel_order_detail(request, order_id):
    order = get_object_or_404(Order, pk=order_id)

    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in dict(Order.STATUS_CHOICES):
            order.status = new_status
            order.save()
            messages.success(request, f'Төлөв шинэчлэгдлээ → {order.get_status_display()}')
        return redirect('core:panel_order_detail', order_id=order_id)

    return render(request, 'panel_order_detail.html', {
        'order':          order,
        'status_choices': Order.STATUS_CHOICES,
        'status_colors':  Order.STATUS_COLORS,
    })
